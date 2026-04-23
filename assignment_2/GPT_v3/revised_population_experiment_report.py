import argparse
from pathlib import Path
import warnings
import time
from typing import Dict, List, Tuple

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import openpyxl
from scipy.optimize import curve_fit, least_squares, minimize

warnings.filterwarnings('ignore')


# =============================================================================
# 字体处理：尽量兼容不同机器，避免中文乱码
# =============================================================================
def configure_fonts(preferred_font=None):
    candidate_fonts = []
    if preferred_font:
        candidate_fonts.append(preferred_font)
    candidate_fonts += [
        'SimHei', 'Microsoft YaHei', 'Noto Sans CJK SC', 'Source Han Sans SC',
        'WenQuanYi Zen Hei', 'PingFang SC', 'Heiti SC', 'Arial Unicode MS', 'DejaVu Sans'
    ]
    available = {f.name for f in fm.fontManager.ttflist}
    selected = None
    for name in candidate_fonts:
        if name in available:
            selected = name
            break
    if selected is None:
        selected = 'DejaVu Sans'
    plt.rcParams['font.sans-serif'] = [selected]
    plt.rcParams['axes.unicode_minus'] = False
    return selected


FONT_NAME = configure_fonts()


# =============================================================================
# 基础指标
# =============================================================================
def mse(y_true, y_pred):
    return float(np.mean((np.asarray(y_true) - np.asarray(y_pred)) ** 2))


def rmse(y_true, y_pred):
    return float(np.sqrt(mse(y_true, y_pred)))


def mae(y_true, y_pred):
    return float(np.mean(np.abs(np.asarray(y_true) - np.asarray(y_pred))))


def mape(y_true, y_pred):
    y_true = np.asarray(y_true)
    y_pred = np.asarray(y_pred)
    eps = 1e-12
    return float(np.mean(np.abs((y_true - y_pred) / np.maximum(np.abs(y_true), eps))) * 100)


def r_squared(y_true, y_pred):
    y_true = np.asarray(y_true)
    y_pred = np.asarray(y_pred)
    ss_res = np.sum((y_true - y_pred) ** 2)
    ss_tot = np.sum((y_true - np.mean(y_true)) ** 2)
    if ss_tot == 0:
        return 1.0
    return float(1 - ss_res / ss_tot)


def ensure_dir(path: Path):
    path.mkdir(parents=True, exist_ok=True)


# =============================================================================
# 数据读取
# =============================================================================
def load_data(excel_path: str, sheet_name: str = 'Sheet1', min_row: int = 6):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'工作表 {sheet_name} 不存在，可用工作表：{wb.sheetnames}')

    sheet = wb[sheet_name]
    years, populations = [], []

    for row in sheet.iter_rows(min_row=min_row, values_only=True):
        if len(row) < 2:
            continue
        year, pop = row[0], row[1]
        if year is None or pop is None:
            continue
        if not isinstance(year, (int, float)):
            continue
        try:
            year = int(year)
            if isinstance(pop, str):
                pop = float(pop.replace('\xa0', '').replace(',', '').strip())
            else:
                pop = float(pop)
        except Exception:
            continue
        years.append(year)
        populations.append(pop)

    if not years:
        raise ValueError('未读取到有效数据，请检查 Excel 文件格式。')

    years = np.array(years, dtype=int)
    populations = np.array(populations, dtype=float) / 10000.0  # 单位：亿人
    return years, populations


# =============================================================================
# 模型定义
# =============================================================================
class PopulationModel:
    name = ''
    formula = ''
    param_names = []

    @staticmethod
    def model_func(t, *params):
        raise NotImplementedError

    @staticmethod
    def initial_guess():
        raise NotImplementedError

    @staticmethod
    def bounds():
        raise NotImplementedError

    @staticmethod
    def is_parameter_valid(params):
        return True, ''


class ExponentialModel(PopulationModel):
    name = '指数增长模型'
    formula = 'N(t)=N0·exp(r·t)'
    param_names = ['N0', 'r']

    @staticmethod
    def model_func(t, N0, r):
        return N0 * np.exp(r * t)

    @staticmethod
    def initial_guess():
        return [5.4, 0.012]

    @staticmethod
    def bounds():
        return ([1.0, -0.05], [20.0, 0.08])

    @staticmethod
    def is_parameter_valid(params):
        N0, _ = params
        if N0 <= 0:
            return False, 'N0 必须大于 0'
        return True, ''


class ImprovedExponentialModel(PopulationModel):
    name = '改进指数增长模型'
    formula = 'N(t)=N0·exp(r·t)·(1+b·t)'
    param_names = ['N0', 'r', 'b']

    @staticmethod
    def model_func(t, N0, r, b):
        return N0 * np.exp(r * t) * (1 + b * t)

    @staticmethod
    def initial_guess():
        return [5.4, 0.012, -0.0005]

    @staticmethod
    def bounds():
        return ([1.0, -0.05, -0.02], [20.0, 0.08, 0.02])

    @staticmethod
    def is_parameter_valid(params):
        N0, _, _ = params
        if N0 <= 0:
            return False, 'N0 必须大于 0'
        return True, ''


class LogisticModel(PopulationModel):
    name = 'Logistic增长模型'
    formula = 'N(t)=K/[1+(K/N0-1)·exp(-r·t)]'
    param_names = ['N0', 'r', 'K']

    @staticmethod
    def model_func(t, N0, r, K):
        return K / (1 + (K / N0 - 1) * np.exp(-r * t))

    @staticmethod
    def initial_guess():
        return [5.4, 0.03, 16.0]

    @staticmethod
    def bounds():
        return ([1.0, -0.05, 8.0], [20.0, 0.10, 30.0])

    @staticmethod
    def is_parameter_valid(params):
        N0, _, K = params
        if N0 <= 0 or K <= 0:
            return False, 'N0 和 K 必须大于 0'
        if K <= N0:
            return False, '应满足 K > N0'
        return True, ''


MODELS = [ExponentialModel, ImprovedExponentialModel, LogisticModel]


# =============================================================================
# 参数估计方法
# =============================================================================
class EstimationMethod:
    name = ''

    @staticmethod
    def estimate(model_class, t, y):
        raise NotImplementedError


class TrustRegionReflective(EstimationMethod):
    name = '信赖域反射法(TRF)'

    @staticmethod
    def estimate(model_class, t, y):
        p0 = model_class.initial_guess()
        bounds = model_class.bounds()
        try:
            popt, _ = curve_fit(
                model_class.model_func,
                t,
                y,
                p0=p0,
                bounds=bounds,
                method='trf',
                maxfev=50000,
            )
            return {'success': True, 'params': popt, 'message': '成功'}
        except Exception as e:
            return {'success': False, 'params': None, 'message': str(e)}


class DogboxMethod(EstimationMethod):
    name = 'Dogbox信赖域法'

    @staticmethod
    def estimate(model_class, t, y):
        p0 = model_class.initial_guess()
        bounds = model_class.bounds()

        def residuals(params):
            return y - model_class.model_func(t, *params)

        try:
            result = least_squares(
                residuals,
                p0,
                bounds=bounds,
                method='dogbox',
                ftol=1e-12,
                xtol=1e-12,
                gtol=1e-12,
                max_nfev=50000,
            )
            return {
                'success': bool(result.success),
                'params': result.x if result.success else None,
                'message': str(result.message),
            }
        except Exception as e:
            return {'success': False, 'params': None, 'message': str(e)}


class LBFGSBMethod(EstimationMethod):
    name = 'L-BFGS-B拟牛顿法'

    @staticmethod
    def estimate(model_class, t, y):
        p0 = model_class.initial_guess()
        bounds = model_class.bounds()

        def objective(params):
            pred = model_class.model_func(t, *params)
            if np.any(~np.isfinite(pred)):
                return 1e20
            return np.sum((y - pred) ** 2)

        try:
            result = minimize(
                objective,
                p0,
                method='L-BFGS-B',
                bounds=list(zip(bounds[0], bounds[1])),
                options={'maxiter': 50000, 'ftol': 1e-12},
            )
            return {
                'success': bool(result.success),
                'params': result.x if result.success else None,
                'message': str(result.message),
            }
        except Exception as e:
            return {'success': False, 'params': None, 'message': str(e)}


METHODS = [TrustRegionReflective, DogboxMethod, LBFGSBMethod]


# =============================================================================
# 实验主逻辑
# =============================================================================
def train_test_split(years, y, train_end_year=2015):
    train_mask = years <= train_end_year
    test_mask = years > train_end_year
    if train_mask.sum() < 5 or test_mask.sum() < 3:
        raise ValueError('训练集或测试集过小，请调整 train_end_year。')
    return years[train_mask], y[train_mask], years[test_mask], y[test_mask]


def evaluate_predictions(y_true, y_pred):
    return {
        'R2': r_squared(y_true, y_pred),
        'RMSE': rmse(y_true, y_pred),
        'MAE': mae(y_true, y_pred),
        'MAPE': mape(y_true, y_pred),
    }


def estimate_and_check(method_class, model_class, t, y, year_grid=None):
    start = time.time()
    est_result = method_class.estimate(model_class, t, y)
    elapsed = time.time() - start

    if (not est_result['success']) or est_result['params'] is None:
        return {'success': False, 'params': None, 'time': elapsed, 'message': est_result.get('message', '未知错误')}

    params = np.asarray(est_result['params'], dtype=float)
    if np.any(~np.isfinite(params)):
        return {'success': False, 'params': None, 'time': elapsed, 'message': '参数存在非有限值'}

    valid, reason = model_class.is_parameter_valid(params)
    if not valid:
        return {'success': False, 'params': None, 'time': elapsed, 'message': f'参数校验失败：{reason}'}

    try:
        pred = model_class.model_func(t, *params)
        if np.any(~np.isfinite(pred)) or np.any(pred <= 0):
            return {'success': False, 'params': None, 'time': elapsed, 'message': '模型输出无效'}
        if year_grid is not None:
            pred_grid = model_class.model_func(year_grid, *params)
            if np.any(~np.isfinite(pred_grid)) or np.any(pred_grid <= 0):
                return {'success': False, 'params': None, 'time': elapsed, 'message': '外推输出无效'}
    except Exception as e:
        return {'success': False, 'params': None, 'time': elapsed, 'message': f'预测失败：{e}'}

    return {'success': True, 'params': params, 'time': elapsed, 'message': '成功'}


def run_experiment(years, pop_yi, train_end_year=2015):
    train_years, train_y, test_years, test_y = train_test_split(years, pop_yi, train_end_year)
    t_train = train_years - years[0]
    t_test = test_years - years[0]
    full_grid = np.arange(years[0], 2051) - years[0]

    results = []
    test_pred_tables = []

    for model in MODELS:
        for method in METHODS:
            est = estimate_and_check(method, model, t_train, train_y, year_grid=full_grid)
            if not est['success']:
                results.append({
                    '模型名称': model.name,
                    '参数估计方法': method.name,
                    '是否成功': '否',
                    '失败原因': est['message'],
                    '运行时间(s)': round(est['time'], 4),
                })
                continue

            params = est['params']
            y_train_pred = model.model_func(t_train, *params)
            y_test_pred = model.model_func(t_test, *params)
            train_metrics = evaluate_predictions(train_y, y_train_pred)
            test_metrics = evaluate_predictions(test_y, y_test_pred)

            item = {
                '模型名称': model.name,
                '参数估计方法': method.name,
                '是否成功': '是',
                '失败原因': '',
                '运行时间(s)': round(est['time'], 4),
                '训练集R2': train_metrics['R2'],
                '训练集RMSE': train_metrics['RMSE'],
                '训练集MAE': train_metrics['MAE'],
                '训练集MAPE(%)': train_metrics['MAPE'],
                '测试集R2': test_metrics['R2'],
                '测试集RMSE': test_metrics['RMSE'],
                '测试集MAE': test_metrics['MAE'],
                '测试集MAPE(%)': test_metrics['MAPE'],
                '参数个数': len(params),
                '参数向量': ', '.join([f'{p:.6f}' for p in params]),
                'raw_params': params,
                'model_class': model,
                'method_class': method,
            }
            results.append(item)

            for year, actual, pred in zip(test_years, test_y, y_test_pred):
                test_pred_tables.append({
                    '模型名称': model.name,
                    '参数估计方法': method.name,
                    '年份': int(year),
                    '实际人口(亿人)': float(actual),
                    '预测人口(亿人)': float(pred),
                    '绝对误差': float(abs(actual - pred)),
                })

    return results, test_pred_tables, (train_years, train_y, test_years, test_y)


def choose_best_per_model(results: List[Dict]):
    best = {}
    model_names = [m.name for m in MODELS]
    for model_name in model_names:
        valid = [r for r in results if r.get('是否成功') == '是' and r['模型名称'] == model_name]
        if not valid:
            continue
        valid.sort(key=lambda x: (x['测试集RMSE'], x['测试集MAPE(%)'], -x['测试集R2']))
        best[model_name] = valid[0]
    return best


def refit_best_models(best_results: Dict[str, Dict], years, pop_yi, future_years):
    refit = {}
    t_all = years - years[0]
    for model_name, row in best_results.items():
        model_class = row['model_class']
        method_class = row['method_class']
        est = estimate_and_check(method_class, model_class, t_all, pop_yi,
                                 year_grid=np.arange(years[0], future_years.max() + 1) - years[0])
        if not est['success']:
            refit[model_name] = {'是否成功': '否', '失败原因': est['message']}
            continue
        params = est['params']
        preds = model_class.model_func(future_years - years[0], *params)
        refit[model_name] = {
            '是否成功': '是',
            '模型名称': model_name,
            '参数估计方法': row['参数估计方法'],
            '参数向量': ', '.join([f'{p:.6f}' for p in params]),
            'raw_params': params,
            'future_years': future_years,
            'future_preds': preds,
            'model_class': model_class,
        }
    return refit


def annualized_growth_rate(v0, v1, years_gap):
    if v0 <= 0 or years_gap <= 0:
        return np.nan
    return float(((v1 / v0) ** (1 / years_gap) - 1) * 100)


# =============================================================================
# 导出表格与结论
# =============================================================================
def save_csv(path: Path, rows: List[Dict], columns: List[str]):
    import csv
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({c: row.get(c, '') for c in columns})


def export_tables(output_dir: Path, results: List[Dict], best_results: Dict[str, Dict], test_pred_tables: List[Dict], refit_info: Dict[str, Dict]):
    ensure_dir(output_dir)

    table1_cols = [
        '模型名称', '参数估计方法', '是否成功', '失败原因', '运行时间(s)',
        '训练集R2', '训练集RMSE', '训练集MAE', '训练集MAPE(%)',
        '测试集R2', '测试集RMSE', '测试集MAE', '测试集MAPE(%)',
        '参数个数', '参数向量'
    ]
    save_csv(output_dir / '表1_模型参数估计与拟合检验结果.csv', results, table1_cols)

    table2_rows = []
    for row in best_results.values():
        table2_rows.append({
            '模型名称': row['模型名称'],
            '最佳参数估计方法': row['参数估计方法'],
            '测试集R2': row['测试集R2'],
            '测试集RMSE': row['测试集RMSE'],
            '测试集MAE': row['测试集MAE'],
            '测试集MAPE(%)': row['测试集MAPE(%)'],
            '参数向量': row['参数向量'],
        })
    save_csv(output_dir / '表2_各模型最佳方法与检验结果.csv', table2_rows,
             ['模型名称', '最佳参数估计方法', '测试集R2', '测试集RMSE', '测试集MAE', '测试集MAPE(%)', '参数向量'])

    save_csv(output_dir / '表3_测试集逐年预测结果.csv', test_pred_tables,
             ['模型名称', '参数估计方法', '年份', '实际人口(亿人)', '预测人口(亿人)', '绝对误差'])

    future_rows = []
    for model_name, info in refit_info.items():
        if info.get('是否成功') != '是':
            continue
        row = {
            '模型名称': model_name,
            '用于重估的最佳方法': info['参数估计方法'],
            '参数向量': info['参数向量'],
        }
        for year, pred in zip(info['future_years'], info['future_preds']):
            row[f'{int(year)}年预测人口(亿人)'] = float(pred)
        # 给出年均增长率，避免原来“不同跨度直接比较”的问题
        pairs = list(zip(info['future_years'][:-1], info['future_years'][1:], info['future_preds'][:-1], info['future_preds'][1:]))
        for y0, y1, p0, p1 in pairs:
            row[f'{int(y0)}-{int(y1)}年均增长率(%)'] = annualized_growth_rate(p0, p1, int(y1 - y0))
        future_rows.append(row)

    dynamic_cols = ['模型名称', '用于重估的最佳方法', '参数向量']
    future_year_cols = []
    growth_cols = []
    if future_rows:
        for k in future_rows[0].keys():
            if k.endswith('预测人口(亿人)'):
                future_year_cols.append(k)
            elif k.endswith('年均增长率(%)'):
                growth_cols.append(k)
    save_csv(output_dir / '表4_未来人口预测与年均增长率.csv', future_rows, dynamic_cols + future_year_cols + growth_cols)


def generate_conclusion(output_dir: Path, years, train_years, test_years, best_results, refit_info, font_name):
    report_path = output_dir / '实验结论与结果分析.md'

    ranking = sorted(best_results.values(), key=lambda x: x['测试集RMSE'])
    best_text = ''
    if ranking:
        best_text = f"测试集上综合表现最好的模型为 **{ranking[0]['模型名称']}**，对应的参数估计方法为 **{ranking[0]['参数估计方法']}**，其测试集 RMSE 为 **{ranking[0]['测试集RMSE']:.6f}**，MAPE 为 **{ranking[0]['测试集MAPE(%)']:.4f}%**。"

    compare_lines = []
    for row in ranking:
        compare_lines.append(
            f"- {row['模型名称']}：最佳方法为 {row['参数估计方法']}，测试集 R2={row['测试集R2']:.6f}，RMSE={row['测试集RMSE']:.6f}，MAPE={row['测试集MAPE(%)']:.4f}%"
        )

    future_lines = []
    for model_name, info in refit_info.items():
        if info.get('是否成功') != '是':
            continue
        preds = info['future_preds']
        years_ = info['future_years']
        desc = '，'.join([f"{int(y)}年约为 {p:.2f} 亿人" for y, p in zip(years_, preds)])
        future_lines.append(f"- {model_name}：{desc}")

    with open(report_path, 'w', encoding='utf-8') as f:
        f.write('# 实验结论与结果分析\n\n')
        f.write('## 1. 实验设置说明\n\n')
        f.write(f'- 数据年份范围：{int(years.min())}—{int(years.max())}年\n')
        f.write(f'- 训练集：{int(train_years.min())}—{int(train_years.max())}年\n')
        f.write(f'- 测试集：{int(test_years.min())}—{int(test_years.max())}年\n')
        f.write(f'- 绘图字体自动检测结果：{font_name}\n\n')

        f.write('## 2. 模型检验结果分析\n\n')
        f.write(best_text + '\n\n')
        f.write('\n'.join(compare_lines) + '\n\n')

        f.write('从测试集误差指标来看，三种模型在中国人口数据上的泛化能力存在明显差异。一般而言，若某模型在训练集上拟合较好，但在测试集上误差明显增大，则说明该模型的外推稳定性较弱；反之，则说明该模型在刻画人口增长趋势方面更稳健。\n\n')

        f.write('## 3. 未来人口预测结果分析\n\n')
        f.write('\n'.join(future_lines) + '\n\n')
        f.write('在未来预测部分，代码统一给出 2024、2025、2030、2040、2050 年的人口预测值，并额外给出分阶段的**年均增长率**，避免把不同时间跨度下的累计增幅直接当作同一种增长率进行比较。\n\n')

        f.write('## 4. 综合结论\n\n')
        f.write('1. 三种模型均可用于中国人口增长趋势的刻画，但其适用范围和预测稳定性不同。\n')
        f.write('2. 指数增长模型结构最简单，但长期预测时往往更容易出现偏离现实的持续增长。\n')
        f.write('3. 改进指数增长模型在一定程度上缓解了简单指数模型的不足，但其修正项形式仍属于经验性设定。\n')
        f.write('4. Logistic 增长模型考虑了人口增长受资源与环境约束的影响，通常更适合长期趋势分析。\n')
        f.write('5. 最终应结合测试集检验指标与未来预测合理性，选取更适合作为报告主结论的模型。\n')


# =============================================================================
# 绘图：更贴近实验报告排版，不再使用有歧义的图
# =============================================================================
def plot_fit_report_style(output_dir: Path, years, pop_yi, train_years, test_years, refit_info):
    valid_items = [(k, v) for k, v in refit_info.items() if v.get('是否成功') == '是']
    if not valid_items:
        return
    fig, ax = plt.subplots(figsize=(11, 7))

    train_mask = np.isin(years, train_years)
    test_mask = np.isin(years, test_years)
    ax.scatter(years[train_mask], pop_yi[train_mask], s=35, marker='o', label='训练集实际数据', zorder=5)
    ax.scatter(years[test_mask], pop_yi[test_mask], s=40, marker='s', label='测试集实际数据', zorder=5)

    draw_years = np.arange(years.min(), 2051)
    t_draw = draw_years - years.min()
    colors = ['#d62728', '#1f77b4', '#2ca02c']

    idx = 0
    for model_name, info in valid_items:
        model_class = info['model_class']
        params = info['raw_params']
        y_draw = model_class.model_func(t_draw, *params)
        hist_mask = draw_years <= years.max()
        future_mask = draw_years > years.max()
        color = colors[idx % len(colors)]
        idx += 1

        ax.plot(draw_years[hist_mask], y_draw[hist_mask], linewidth=2.2, color=color,
                label=f"{model_name}（{info['参数估计方法']}）")
        ax.plot(draw_years[future_mask], y_draw[future_mask], linewidth=2.2, linestyle='--', color=color)

    ax.axvline(x=train_years.max(), linestyle='--', linewidth=1.1, color='gray', label='训练集/测试集分界')
    ax.axvline(x=years.max(), linestyle=':', linewidth=1.1, color='gray', label='历史数据/预测区间分界')
    ax.set_title('图1 三种人口增长模型拟合结果与未来预测曲线')
    ax.set_xlabel('年份')
    ax.set_ylabel('人口（亿人）')
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=9)
    plt.tight_layout()
    plt.savefig(output_dir / '图1_三种人口增长模型拟合结果与未来预测曲线.png', dpi=180, bbox_inches='tight')
    plt.close()


def plot_test_metric_report_style(output_dir: Path, best_results: Dict[str, Dict]):
    if not best_results:
        return
    models = list(best_results.keys())
    rmse_vals = [best_results[m]['测试集RMSE'] for m in models]
    mape_vals = [best_results[m]['测试集MAPE(%)'] for m in models]

    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    x = np.arange(len(models))

    bars1 = axes[0].bar(x, rmse_vals, width=0.55)
    axes[0].set_xticks(x)
    axes[0].set_xticklabels(models, rotation=10)
    axes[0].set_ylabel('RMSE')
    axes[0].set_title('图2(a) 测试集RMSE比较')
    axes[0].grid(True, alpha=0.3, axis='y')
    for bar, v in zip(bars1, rmse_vals):
        axes[0].text(bar.get_x() + bar.get_width() / 2, bar.get_height(), f'{v:.4f}', ha='center', va='bottom', fontsize=9)

    bars2 = axes[1].bar(x, mape_vals, width=0.55)
    axes[1].set_xticks(x)
    axes[1].set_xticklabels(models, rotation=10)
    axes[1].set_ylabel('MAPE（%）')
    axes[1].set_title('图2(b) 测试集MAPE比较')
    axes[1].grid(True, alpha=0.3, axis='y')
    for bar, v in zip(bars2, mape_vals):
        axes[1].text(bar.get_x() + bar.get_width() / 2, bar.get_height(), f'{v:.2f}%', ha='center', va='bottom', fontsize=9)

    plt.tight_layout()
    plt.savefig(output_dir / '图2_测试集误差指标比较.png', dpi=180, bbox_inches='tight')
    plt.close()


def plot_future_prediction_report_style(output_dir: Path, years, pop_yi, refit_info):
    valid_items = [(k, v) for k, v in refit_info.items() if v.get('是否成功') == '是']
    if not valid_items:
        return
    fig, ax = plt.subplots(figsize=(10, 6))
    last_year = int(years.max())
    last_pop = float(pop_yi[-1])
    ax.scatter([last_year], [last_pop], s=55, label=f'{last_year}年实际人口', zorder=5)

    colors = ['#d62728', '#1f77b4', '#2ca02c']
    idx = 0
    for model_name, info in valid_items:
        years_ = info['future_years']
        preds = info['future_preds']
        x = np.concatenate([[last_year], years_])
        y = np.concatenate([[last_pop], preds])
        ax.plot(x, y, marker='o', linewidth=2.2, color=colors[idx % len(colors)],
                label=f"{model_name}（{info['参数估计方法']}）")
        idx += 1

    ax.set_title('图3 三种人口增长模型未来人口预测对比')
    ax.set_xlabel('年份')
    ax.set_ylabel('人口（亿人）')
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=9)
    plt.tight_layout()
    plt.savefig(output_dir / '图3_三种人口增长模型未来人口预测对比.png', dpi=180, bbox_inches='tight')
    plt.close()


def plot_test_prediction_comparison(output_dir: Path, years, pop_yi, train_years, test_years, best_results):
    if not best_results:
        return
    fig, ax = plt.subplots(figsize=(10, 6))

    test_mask = np.isin(years, test_years)
    ax.plot(test_years, pop_yi[test_mask], 'o-', linewidth=2.0, label='测试集实际人口')

    colors = ['#d62728', '#1f77b4', '#2ca02c']
    for idx, (model_name, row) in enumerate(best_results.items()):
        model_class = row['model_class']
        params = row['raw_params']
        y_pred = model_class.model_func(test_years - years.min(), *params)
        ax.plot(test_years, y_pred, marker='s', linewidth=1.8, linestyle='--', color=colors[idx % len(colors)],
                label=f"{model_name}预测")

    ax.set_title('图4 测试集实际值与最佳模型预测值对比')
    ax.set_xlabel('年份')
    ax.set_ylabel('人口（亿人）')
    ax.grid(True, alpha=0.3)
    ax.legend(fontsize=9)
    plt.tight_layout()
    plt.savefig(output_dir / '图4_测试集实际值与最佳模型预测值对比.png', dpi=180, bbox_inches='tight')
    plt.close()


# =============================================================================
# 主程序
# =============================================================================
def main():
    parser = argparse.ArgumentParser(description='中国人口增长模型分析实验（报告版）')
    parser.add_argument('--excel', type=str, required=True, help='Excel 数据文件路径')
    parser.add_argument('--sheet', type=str, default='Sheet1', help='工作表名称')
    parser.add_argument('--min-row', type=int, default=6, help='数据起始行')
    parser.add_argument('--train-end-year', type=int, default=2015, help='训练集截止年份')
    parser.add_argument('--output-dir', type=str, default='population_report_outputs', help='输出目录')
    parser.add_argument('--font', type=str, default='', help='可选：手动指定绘图字体名称')
    args = parser.parse_args()

    global FONT_NAME
    if args.font:
        FONT_NAME = configure_fonts(args.font)

    output_dir = Path(args.output_dir)
    ensure_dir(output_dir)

    years, pop_yi = load_data(args.excel, args.sheet, args.min_row)
    results, test_pred_tables, split_data = run_experiment(years, pop_yi, args.train_end_year)
    train_years, train_y, test_years, test_y = split_data

    best_results = choose_best_per_model(results)
    future_years = np.array([2024, 2025, 2030, 2040, 2050], dtype=int)
    refit_info = refit_best_models(best_results, years, pop_yi, future_years)

    export_tables(output_dir, results, best_results, test_pred_tables, refit_info)
    generate_conclusion(output_dir, years, train_years, test_years, best_results, refit_info, FONT_NAME)

    plot_fit_report_style(output_dir, years, pop_yi, train_years, test_years, refit_info)
    plot_test_metric_report_style(output_dir, best_results)
    plot_future_prediction_report_style(output_dir, years, pop_yi, refit_info)
    plot_test_prediction_comparison(output_dir, years, pop_yi, train_years, test_years, best_results)

    print('=' * 72)
    print('中国人口增长模型分析实验（报告版）运行完成')
    print('=' * 72)
    print(f'数据年份范围：{int(years.min())}—{int(years.max())}年')
    print(f'训练集：{int(train_years.min())}—{int(train_years.max())}年；测试集：{int(test_years.min())}—{int(test_years.max())}年')
    print(f'当前绘图字体：{FONT_NAME}')
    print(f'输出目录：{output_dir.resolve()}')
    print('\n各模型最佳结果：')
    for model_name, row in best_results.items():
        print(f"- {model_name}：{row['参数估计方法']}，测试集RMSE={row['测试集RMSE']:.6f}，MAPE={row['测试集MAPE(%)']:.4f}%")


if __name__ == '__main__':
    main()
