from __future__ import annotations

"""
人口增长模型分析实验（修订版）

相对原脚本，主要修正：
1. 保留三种必选模型，并支持两个扩展模型。
2. 增加训练集/检验集划分，避免“用全部数据拟合再在同数据上检验”。
3. 增加未来年份预测表。
4. 修正部分方法命名与实现不一致的问题。
5. 去除硬编码路径，改为命令行参数/相对路径。
6. 导出汇总表与图片，便于写实验报告。
"""

import argparse
import json
import math
import time
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, List, Optional, Sequence, Tuple

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from scipy.optimize import curve_fit, least_squares, minimize

warnings.filterwarnings("ignore")
plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "Arial Unicode MS", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False


# =========================
# 数据读取
# =========================
def load_population_data(excel_path: Path, sheet_name: str = "Sheet1", start_row: int = 6) -> Tuple[np.ndarray, np.ndarray]:
    """从 Excel 读取年份和人口数据。

    约定：
    - 第 1 列为年份
    - 第 2 列为人口（单位：万人）
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"未找到数据文件: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表 {sheet_name!r} 不存在，可选：{wb.sheetnames}")
    sheet = wb[sheet_name]

    years: List[int] = []
    populations: List[float] = []
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        if len(row) < 2:
            continue
        year, pop = row[0], row[1]
        if year is None or pop is None:
            continue
        if not isinstance(year, (int, float)):
            continue
        if isinstance(pop, str):
            pop = float(str(pop).replace("\xa0", "").replace(",", "").strip())
        years.append(int(year))
        populations.append(float(pop))

    if not years:
        raise ValueError("未读取到有效数据，请检查起始行、工作表名和列位置。")

    return np.array(years, dtype=float), np.array(populations, dtype=float)


# =========================
# 评价指标
# =========================
def mse(y_true: np.ndarray, y_pred: np.ndarray) -> float:
    return float(np.mean((y_true - y_pred) ** 2))


def rmse(y_true: np.ndarray, y_pred: np.ndarray) -> float:
    return float(np.sqrt(mse(y_true, y_pred)))


def mae(y_true: np.ndarray, y_pred: np.ndarray) -> float:
    return float(np.mean(np.abs(y_true - y_pred)))


def mape(y_true: np.ndarray, y_pred: np.ndarray) -> float:
    denom = np.where(np.abs(y_true) < 1e-12, 1e-12, np.abs(y_true))
    return float(np.mean(np.abs((y_true - y_pred) / denom)) * 100)


def r_squared(y_true: np.ndarray, y_pred: np.ndarray) -> float:
    ss_res = float(np.sum((y_true - y_pred) ** 2))
    ss_tot = float(np.sum((y_true - np.mean(y_true)) ** 2))
    if ss_tot < 1e-12:
        return 1.0
    return 1 - ss_res / ss_tot


def build_metrics(y_true: np.ndarray, y_pred: np.ndarray) -> Dict[str, float]:
    return {
        "R2": r_squared(y_true, y_pred),
        "MSE": mse(y_true, y_pred),
        "RMSE": rmse(y_true, y_pred),
        "MAE": mae(y_true, y_pred),
        "MAPE(%)": mape(y_true, y_pred),
    }


# =========================
# 模型定义
# =========================
class PopulationModel:
    name = ""
    formula = ""
    param_names: Sequence[str] = ()

    @staticmethod
    def model_func(t: np.ndarray, *params: float) -> np.ndarray:
        raise NotImplementedError

    @staticmethod
    def initial_guess() -> List[float]:
        raise NotImplementedError

    @staticmethod
    def bounds() -> Tuple[Sequence[float], Sequence[float]]:
        raise NotImplementedError


class ExponentialModel(PopulationModel):
    name = "指数增长模型"
    formula = "N(t)=N0*exp(r*t)"
    param_names = ("N0", "r")

    @staticmethod
    def model_func(t, N0, r):
        return N0 * np.exp(r * t)

    @staticmethod
    def initial_guess():
        return [5.4, 0.015]

    @staticmethod
    def bounds():
        return ([1.0, -0.05], [30.0, 0.1])


class ImprovedExponentialModel(PopulationModel):
    name = "改进指数模型"
    formula = "N(t)=N0*exp(r*t)*(1+b*t)"
    param_names = ("N0", "r", "b")

    @staticmethod
    def model_func(t, N0, r, b):
        return N0 * np.exp(r * t) * (1 + b * t)

    @staticmethod
    def initial_guess():
        return [5.4, 0.012, -0.0002]

    @staticmethod
    def bounds():
        return ([1.0, -0.05, -0.02], [30.0, 0.1, 0.02])


class LogisticModel(PopulationModel):
    name = "Logistic增长模型"
    formula = "N(t)=K/(1+(K/N0-1)*exp(-r*t))"
    param_names = ("N0", "r", "K")

    @staticmethod
    def model_func(t, N0, r, K):
        return K / (1 + (K / np.maximum(N0, 1e-12) - 1) * np.exp(-r * t))

    @staticmethod
    def initial_guess():
        return [5.4, 0.03, 16.0]

    @staticmethod
    def bounds():
        return ([1.0, -0.05, 10.0], [30.0, 0.2, 40.0])


class GompertzModel(PopulationModel):
    name = "Gompertz增长模型"
    formula = "N(t)=K*exp(-b*exp(-r*t))"
    param_names = ("K", "b", "r")

    @staticmethod
    def model_func(t, K, b, r):
        return K * np.exp(-b * np.exp(-r * t))

    @staticmethod
    def initial_guess():
        return [15.0, 3.0, 0.05]

    @staticmethod
    def bounds():
        return ([5.0, 0.01, -0.05], [40.0, 20.0, 0.2])


class VonBertalanffyModel(PopulationModel):
    name = "Von Bertalanffy增长模型"
    formula = "N(t)=K*(1-exp(-r*(t-t0)))^3"
    param_names = ("K", "r", "t0")

    @staticmethod
    def model_func(t, K, r, t0):
        diff = 1 - np.exp(-r * (t - t0))
        return K * np.power(np.maximum(diff, 1e-10), 3)

    @staticmethod
    def initial_guess():
        return [20.0, 0.03, -5.0]

    @staticmethod
    def bounds():
        return ([5.0, 0.001, -50.0], [50.0, 0.2, 50.0])


REQUIRED_MODELS = [ExponentialModel, ImprovedExponentialModel, LogisticModel]
EXTRA_MODELS = [GompertzModel, VonBertalanffyModel]


# =========================
# 参数估计方法
# =========================
class EstimationMethod:
    name = ""

    @staticmethod
    def estimate(model_class: type[PopulationModel], t: np.ndarray, y: np.ndarray) -> Dict[str, object]:
        raise NotImplementedError


class CurveFitTRFMethod(EstimationMethod):
    name = "非线性最小二乘法(curve_fit-TRF)"

    @staticmethod
    def estimate(model_class, t, y):
        try:
            p0 = model_class.initial_guess()
            bounds = model_class.bounds()
            popt, _ = curve_fit(
                model_class.model_func,
                t,
                y,
                p0=p0,
                bounds=bounds,
                method="trf",
                maxfev=50000,
            )
            return {"params": np.array(popt, dtype=float), "success": True, "message": "成功"}
        except Exception as e:
            return {"params": None, "success": False, "message": str(e)}


class LeastSquaresDogboxMethod(EstimationMethod):
    name = "有界最小二乘法(least_squares-dogbox)"

    @staticmethod
    def estimate(model_class, t, y):
        try:
            p0 = model_class.initial_guess()
            bounds = model_class.bounds()

            def residuals(params):
                return y - model_class.model_func(t, *params)

            result = least_squares(
                residuals,
                p0,
                bounds=bounds,
                method="dogbox",
                ftol=1e-10,
                xtol=1e-10,
                gtol=1e-10,
                max_nfev=50000,
            )
            return {
                "params": result.x if result.success else None,
                "success": bool(result.success),
                "message": str(result.message),
            }
        except Exception as e:
            return {"params": None, "success": False, "message": str(e)}


class GradientDescentMethod(EstimationMethod):
    name = "梯度型优化(L-BFGS-B)"

    @staticmethod
    def estimate(model_class, t, y):
        try:
            p0 = np.array(model_class.initial_guess(), dtype=float)
            bounds = model_class.bounds()

            def objective(params):
                pred = model_class.model_func(t, *params)
                return float(np.sum((y - pred) ** 2))

            result = minimize(
                objective,
                p0,
                method="L-BFGS-B",
                bounds=list(zip(bounds[0], bounds[1])),
                options={"maxiter": 50000, "ftol": 1e-12},
            )
            return {
                "params": result.x if result.success else None,
                "success": bool(result.success),
                "message": str(result.message),
            }
        except Exception as e:
            return {"params": None, "success": False, "message": str(e)}


class TrustRegionMethod(EstimationMethod):
    name = "信赖域反射法(least_squares-TRF)"

    @staticmethod
    def estimate(model_class, t, y):
        try:
            p0 = model_class.initial_guess()
            bounds = model_class.bounds()

            def residuals(params):
                return y - model_class.model_func(t, *params)

            result = least_squares(
                residuals,
                p0,
                bounds=bounds,
                method="trf",
                ftol=1e-12,
                xtol=1e-12,
                gtol=1e-12,
                max_nfev=50000,
            )
            return {
                "params": result.x if result.success else None,
                "success": bool(result.success),
                "message": str(result.message),
            }
        except Exception as e:
            return {"params": None, "success": False, "message": str(e)}


METHODS = [CurveFitTRFMethod, LeastSquaresDogboxMethod, GradientDescentMethod, TrustRegionMethod]


@dataclass
class ExperimentResult:
    model_name: str
    method_name: str
    success: bool
    params: Optional[np.ndarray]
    fit_metrics: Optional[Dict[str, float]]
    test_metrics: Optional[Dict[str, float]]
    elapsed_seconds: float
    message: str


def split_train_test(years: np.ndarray, y: np.ndarray, test_size: int) -> Dict[str, np.ndarray]:
    if test_size <= 0 or test_size >= len(years):
        raise ValueError("test_size 必须大于 0 且小于样本总数")

    split_idx = len(years) - test_size
    return {
        "train_years": years[:split_idx],
        "test_years": years[split_idx:],
        "train_y": y[:split_idx],
        "test_y": y[split_idx:],
    }


def format_params(model_cls: type[PopulationModel], params: np.ndarray) -> str:
    return ", ".join(f"{name}={value:.6f}" for name, value in zip(model_cls.param_names, params))


def run_experiment(
    models: Sequence[type[PopulationModel]],
    methods: Sequence[type[EstimationMethod]],
    years: np.ndarray,
    y: np.ndarray,
    test_size: int,
) -> Tuple[List[ExperimentResult], Dict[str, np.ndarray]]:
    split = split_train_test(years, y, test_size=test_size)
    t0 = years[0]
    train_t = split["train_years"] - t0
    test_t = split["test_years"] - t0
    all_t = years - t0

    results: List[ExperimentResult] = []

    for method in methods:
        print("\n" + "=" * 72)
        print(f"参数估计方法：{method.name}")
        print("=" * 72)

        for model in models:
            start = time.time()
            est_result = method.estimate(model, train_t, split["train_y"])
            elapsed = time.time() - start

            if est_result["success"] and est_result["params"] is not None:
                params = np.array(est_result["params"], dtype=float)
                all_pred = model.model_func(all_t, *params)
                test_pred = model.model_func(test_t, *params)

                fit_metrics = build_metrics(y, all_pred)
                test_metrics = build_metrics(split["test_y"], test_pred)

                results.append(
                    ExperimentResult(
                        model_name=model.name,
                        method_name=method.name,
                        success=True,
                        params=params,
                        fit_metrics=fit_metrics,
                        test_metrics=test_metrics,
                        elapsed_seconds=elapsed,
                        message="成功",
                    )
                )

                print(f"\n{model.name}")
                print(f"公式: {model.formula}")
                print(f"参数: {format_params(model, params)}")
                print(f"全样本拟合: R2={fit_metrics['R2']:.6f}, RMSE={fit_metrics['RMSE']:.6f}, MAPE={fit_metrics['MAPE(%)']:.3f}%")
                print(f"检验集表现: RMSE={test_metrics['RMSE']:.6f}, MAPE={test_metrics['MAPE(%)']:.3f}%")
                print(f"耗时: {elapsed:.4f}s")
            else:
                results.append(
                    ExperimentResult(
                        model_name=model.name,
                        method_name=method.name,
                        success=False,
                        params=None,
                        fit_metrics=None,
                        test_metrics=None,
                        elapsed_seconds=elapsed,
                        message=str(est_result["message"]),
                    )
                )
                print(f"\n{model.name}: 估计失败 - {est_result['message']}")

    return results, split


def build_summary_dataframe(results: Sequence[ExperimentResult]) -> pd.DataFrame:
    rows = []
    for res in results:
        row = {
            "方法": res.method_name,
            "模型": res.model_name,
            "是否成功": res.success,
            "耗时(s)": round(res.elapsed_seconds, 4),
            "消息": res.message,
        }
        if res.success and res.params is not None and res.fit_metrics and res.test_metrics:
            row.update({
                "参数": json.dumps([round(float(x), 6) for x in res.params], ensure_ascii=False),
                "拟合R2": res.fit_metrics["R2"],
                "拟合RMSE": res.fit_metrics["RMSE"],
                "拟合MAE": res.fit_metrics["MAE"],
                "拟合MAPE(%)": res.fit_metrics["MAPE(%)"],
                "检验RMSE": res.test_metrics["RMSE"],
                "检验MAE": res.test_metrics["MAE"],
                "检验MAPE(%)": res.test_metrics["MAPE(%)"],
            })
        rows.append(row)
    df = pd.DataFrame(rows)
    if not df.empty and "检验RMSE" in df.columns:
        df = df.sort_values(by=["是否成功", "检验RMSE"], ascending=[False, True])
    return df


def pick_best_results(results: Sequence[ExperimentResult]) -> Dict[str, ExperimentResult]:
    best: Dict[str, ExperimentResult] = {}
    for res in results:
        if not res.success or not res.test_metrics:
            continue
        old = best.get(res.model_name)
        if old is None or res.test_metrics["RMSE"] < old.test_metrics["RMSE"]:
            best[res.model_name] = res
    return best


def build_prediction_dataframe(
    best_by_model: Dict[str, ExperimentResult],
    years: np.ndarray,
    future_years: Sequence[int],
    base_year: int,
) -> pd.DataFrame:
    rows = []
    year_values = list(map(int, years)) + list(map(int, future_years))
    year_values = sorted(set(year_values))
    t = np.array(year_values, dtype=float) - base_year

    for year, tt in zip(year_values, t):
        row = {"年份": year}
        for model_name, res in best_by_model.items():
            model_cls = next(m for m in REQUIRED_MODELS + EXTRA_MODELS if m.name == model_name)
            row[model_name] = float(model_cls.model_func(np.array([tt]), *res.params)[0])
        rows.append(row)

    return pd.DataFrame(rows)


def plot_best_models(
    best_by_model: Dict[str, ExperimentResult],
    years: np.ndarray,
    y: np.ndarray,
    split: Dict[str, np.ndarray],
    output_file: Path,
) -> None:
    if not best_by_model:
        return

    base_year = int(years[0])
    future_end = max(int(years[-1]) + 15, int(split["test_years"][-1]) + 10)
    smooth_years = np.linspace(base_year, future_end, 400)
    smooth_t = smooth_years - base_year

    plt.figure(figsize=(12, 7))
    plt.scatter(split["train_years"], split["train_y"], label="训练集", s=35)
    plt.scatter(split["test_years"], split["test_y"], label="检验集", s=45, marker="s")

    for model_name, res in best_by_model.items():
        model_cls = next(m for m in REQUIRED_MODELS + EXTRA_MODELS if m.name == model_name)
        pred = model_cls.model_func(smooth_t, *res.params)
        label = f"{model_name} | {res.method_name}"
        plt.plot(smooth_years, pred, linewidth=2, label=label)

    plt.title("各模型最优结果对比（含检验集与外推曲线）")
    plt.xlabel("年份")
    plt.ylabel("人口（亿人）")
    plt.grid(True, alpha=0.3)
    plt.legend(fontsize=8)
    plt.tight_layout()
    plt.savefig(output_file, dpi=180, bbox_inches="tight")
    plt.close()


def write_report_markdown(
    summary_df: pd.DataFrame,
    best_by_model: Dict[str, ExperimentResult],
    split: Dict[str, np.ndarray],
    prediction_df: pd.DataFrame,
    output_file: Path,
) -> None:
    lines = []
    lines.append("# 人口增长模型实验结果摘要\n")
    lines.append(f"- 训练集年份：{int(split['train_years'][0])}–{int(split['train_years'][-1])}")
    lines.append(f"- 检验集年份：{int(split['test_years'][0])}–{int(split['test_years'][-1])}")
    lines.append("")
    lines.append("## 1. 各模型-方法组合结果（按检验RMSE排序）\n")
    lines.append(summary_df.to_markdown(index=False))
    lines.append("\n## 2. 每个模型的最优方法\n")
    for model_name, res in best_by_model.items():
        lines.append(
            f"- **{model_name}**：{res.method_name}，检验RMSE={res.test_metrics['RMSE']:.6f}，"
            f"检验MAPE={res.test_metrics['MAPE(%)']:.3f}%"
        )
    lines.append("\n## 3. 未来年份预测（单位：亿人）\n")
    lines.append(prediction_df.to_markdown(index=False, floatfmt='.4f'))
    output_file.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="人口增长模型实验（修订版）")
    parser.add_argument("--excel", type=str, default="1949-2023人口数据-实验1.xlsx", help="人口数据 Excel 文件路径")
    parser.add_argument("--sheet", type=str, default="Sheet1", help="Excel 工作表名")
    parser.add_argument("--start-row", type=int, default=6, help="数据起始行")
    parser.add_argument("--test-size", type=int, default=8, help="末尾留作检验集的数据点个数")
    parser.add_argument(
        "--future-years",
        type=int,
        nargs="*",
        default=[2025, 2030, 2035, 2040, 2050],
        help="需要预测的未来年份",
    )
    parser.add_argument("--output-dir", type=str, default="population_experiment_outputs", help="输出目录")
    parser.add_argument("--use-extra-models", action="store_true", help="是否加入两个扩展模型")
    args = parser.parse_args()

    excel_path = Path(args.excel).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    years, populations_wan = load_population_data(excel_path, sheet_name=args.sheet, start_row=args.start_row)
    populations_yi = populations_wan / 10000.0

    models = REQUIRED_MODELS + EXTRA_MODELS if args.use_extra_models else REQUIRED_MODELS

    print("=" * 84)
    print("人口增长模型实验（修订版）")
    print("=" * 84)
    print(f"数据文件: {excel_path}")
    print(f"年份范围: {int(years[0])}-{int(years[-1])}")
    print(f"样本数: {len(years)}")
    print(f"人口范围: {populations_yi.min():.3f}亿 - {populations_yi.max():.3f}亿")
    print(f"检验集大小: {args.test_size}")
    print(f"模型数: {len(models)}，方法数: {len(METHODS)}")

    results, split = run_experiment(models, METHODS, years, populations_yi, args.test_size)
    summary_df = build_summary_dataframe(results)
    best_by_model = pick_best_results(results)
    prediction_df = build_prediction_dataframe(best_by_model, years, args.future_years, int(years[0]))

    summary_path = output_dir / "实验结果汇总.csv"
    prediction_path = output_dir / "未来年份预测表.csv"
    plot_path = output_dir / "最优模型对比图.png"
    report_path = output_dir / "实验结果摘要.md"

    summary_df.to_csv(summary_path, index=False, encoding="utf-8-sig")
    prediction_df.to_csv(prediction_path, index=False, encoding="utf-8-sig")
    plot_best_models(best_by_model, years, populations_yi, split, plot_path)
    write_report_markdown(summary_df, best_by_model, split, prediction_df, report_path)

    print("\n" + "=" * 84)
    print("输出文件")
    print("=" * 84)
    print(summary_path)
    print(prediction_path)
    print(plot_path)
    print(report_path)

    if not summary_df.empty:
        print("\n按检验RMSE排序的前几项：")
        show_cols = [c for c in ["方法", "模型", "检验RMSE", "检验MAPE(%)", "拟合R2"] if c in summary_df.columns]
        print(summary_df[show_cols].head(10).to_string(index=False))


if __name__ == "__main__":
    main()
